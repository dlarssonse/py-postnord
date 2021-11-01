import json
import time
import os
import random
import sys

import openpyxl
import requests

from pathlib import Path


def main(xlsx_file):
    # Variables for column positions
    col_name = -1
    col_email = -1
    col_address = -1
    col_zip = -1
    col_city = -1

    # Check stuff
    authkey = os.getenv("AUTHKEY")

    if not authkey:
        print("Environment variable AUTHKEY must be set, exiting...")
        exit(1)

    print(f"Importing data from: {xlsx_file}...")

    # Open workbook
    wb = openpyxl.load_workbook(xlsx_file)
    sheet = wb.active

    # Get column position
    for index, col in enumerate(sheet.iter_cols(1, sheet.max_column)):
        if col[0].value == "Namn":
            col_name = index
        if col[0].value == "E-post":
            col_email = index
        if col[0].value == "Adress":
            col_address = index
        if col[0].value == "Postnummer":
            col_zip = index
        if col[0].value == "Ort":
            col_city = index

    # Make dictionary
    data = []
    for i, row in enumerate(sheet.iter_rows(2, values_only=True)):
        data.append(
            {
                "Name": row[col_name],
                "Email": row[col_email],
                "Address": row[col_address],
                "Zipcode": row[col_zip],
                "City": row[col_city],
            }
        )

    # Output persons
    for person in data:
        # Make request
        headers = {
            "Accept": "application/json, text/plain, */*",
            "Accept-Encoding": "gzip, deflate, br",
            "Accept-Language": "sv,en;q=0.9,en-GB;q=0.8,en-US;q=0.7",
            "Authorization": authkey,
            "Content-Type": "application/json;charset=UTF-8",
            "Host": "portal.postnord.com",
            "Origin": "https://portal.postnord.com",
            "Referer": "https://portal.postnord.com/skickadirekt/user/receivers",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.54 Safari/537.36 Edg/95.0.1020.40",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
        }
        pload = {
            "name": person["Name"],
            "email": person["Email"],
            "address": {
                "street": person["Address"],
                "zipCode": person["Zipcode"],
                "city": person["City"],
                "countryCode": "SE",
            },
            "country": {
                "countryCode": "SE",
                "sortIndex": 1,
                "flagUrl": "flags/SE@2x.png",
                "meta": {"euMemberState": True, "udlandZone": "Europa 1"},
                "callingCode": ["46"],
                "currency": "SEK",
                "postalCodeRegExp": "^\\d\\d\\d ?\\d\\d$",
                "postalCodeExample": "11122",
                "name": "Sverige",
            },
        }

        # Send data
        r = requests.post("https://portal.postnord.com/api/receiver/receivers", data=json.dumps(pload), headers=headers)
        if r.status_code == 200:
            print(f"Added '{person['Name']}' ")
        else:
            print(f"Failed to add '{person['Name']}'...")
            print(r.json())

        # Sleep to avoid rate limiting
        time.sleep(random.randint(25, 35))


# Run main
if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python ./postnord.py <FILENAME>")
        exit(1)
    main(sys.argv[1])
